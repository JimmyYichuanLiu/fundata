"""Regression tests for the SQLite-to-Excel fund NAV exporter."""

from __future__ import annotations

import os
import sqlite3
import subprocess
import sys
from contextlib import closing
from pathlib import Path

import pytest
from openpyxl import load_workbook

from organize_fund_data import organize_fund_data


PROJECT_ROOT = Path(__file__).parent.parent


def _create_nav_db(db_path: Path, rows: list[tuple]) -> None:
    with closing(sqlite3.connect(db_path)) as conn:
        conn.execute(
            """
            CREATE TABLE fund_nav_data (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                fund_code TEXT NOT NULL,
                fund_name TEXT,
                nav_date TEXT NOT NULL,
                unit_nav REAL NOT NULL,
                accum_nav REAL,
                data_source TEXT
            )
            """
        )
        conn.executemany(
            """
            INSERT INTO fund_nav_data
                (fund_code, fund_name, nav_date, unit_nav, accum_nav, data_source)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            rows,
        )
        conn.commit()


def _sheet_rows(workbook, sheet_name: str) -> list[tuple]:
    return list(workbook[sheet_name].iter_rows(values_only=True))


def test_exports_only_email_rows_from_migrated_english_schema(tmp_path: Path) -> None:
    db_path = tmp_path / "fund_data.db"
    output_path = tmp_path / "email_nav.xlsx"
    _create_nav_db(
        db_path,
        [
            ("EMAIL01", "邮件基金", "2026-01-01", 1.0, 1.0, "email"),
            ("EMAIL01", "邮件基金", "20260102", 1.1, 1.1, "email"),
            ("ZX01", "臻选基金", "2026-01-01", 2.0, 2.0, "zx_excel"),
        ],
    )

    result = organize_fund_data(db_path, output_path, source="email")

    assert result == output_path.resolve()
    workbook = load_workbook(output_path, read_only=True, data_only=True)
    assert workbook.sheetnames == ["汇总", "EMAIL01"]

    summary_rows = _sheet_rows(workbook, "汇总")
    assert summary_rows[0][:3] == ("产品代码", "产品名称", "记录数")
    assert summary_rows[1][0:3] == ("EMAIL01", "邮件基金", 2)

    detail_rows = _sheet_rows(workbook, "EMAIL01")
    assert detail_rows[0] == ("产品名称", "产品代码", "净值日期", "单位净值", "累计单位净值")
    assert [row[2] for row in detail_rows[1:]] == ["2026-01-01", "2026-01-02"]
    workbook.close()


def test_chinese_headers_and_fund_names_get_readable_column_widths(
    tmp_path: Path,
) -> None:
    db_path = tmp_path / "fund_data.db"
    output_path = tmp_path / "readable_columns.xlsx"
    _create_nav_db(
        db_path,
        [
            (
                "ALU48B",
                "惠晨新秩序2号私募证券投资基金B类",
                "2026-06-26",
                0.9993,
                0.9993,
                "email",
            )
        ],
    )

    organize_fund_data(db_path, output_path, source="email")

    workbook = load_workbook(output_path, read_only=False, data_only=True)
    detail_sheet = workbook["ALU48B"]
    assert detail_sheet.column_dimensions["A"].width >= 36
    assert detail_sheet.column_dimensions["E"].width >= 14
    workbook.close()


def test_mixed_date_formats_are_sorted_chronologically_in_detail_and_summary(tmp_path: Path) -> None:
    db_path = tmp_path / "fund_data.db"
    output_path = tmp_path / "mixed_dates.xlsx"
    _create_nav_db(
        db_path,
        [
            ("MIXED", "混合日期基金", "20260102", 1.02, 1.02, "email"),
            ("MIXED", "混合日期基金", "2026-01-10", 1.10, 1.10, "email"),
            ("MIXED", "混合日期基金", "2025-12-31", 0.99, 0.99, "email"),
        ],
    )

    organize_fund_data(db_path, output_path, source="email")

    workbook = load_workbook(output_path, read_only=True, data_only=True)
    summary = _sheet_rows(workbook, "汇总")[1]
    assert summary[3:7] == ("2025-12-31", "2026-01-10", 0.99, 1.10)

    detail_rows = _sheet_rows(workbook, "MIXED")[1:]
    assert [row[2] for row in detail_rows] == [
        "2025-12-31",
        "2026-01-02",
        "2026-01-10",
    ]
    workbook.close()


def test_invalid_long_and_colliding_fund_codes_get_safe_unique_sheet_names(tmp_path: Path) -> None:
    db_path = tmp_path / "fund_data.db"
    output_path = tmp_path / "safe_sheets.xlsx"
    _create_nav_db(
        db_path,
        [
            ("A/B", "斜杠代码", "2026-01-01", 1.0, None, "email"),
            ("A:B", "冒号代码", "2026-01-01", 1.0, None, "email"),
            ("X" * 40, "超长代码", "2026-01-01", 1.0, None, "email"),
        ],
    )

    organize_fund_data(db_path, output_path, source="email")

    workbook = load_workbook(output_path, read_only=True, data_only=True)
    detail_names = workbook.sheetnames[1:]
    assert len(detail_names) == 3
    assert len({name.casefold() for name in detail_names}) == 3
    assert all(len(name) <= 31 for name in detail_names)
    assert all(not any(char in name for char in "[]:*?/\\") for name in detail_names)

    exported_codes = {
        _sheet_rows(workbook, sheet_name)[1][1]
        for sheet_name in detail_names
    }
    assert exported_codes == {"A/B", "A:B", "X" * 40}
    workbook.close()


def test_source_all_includes_email_and_zx_rows(tmp_path: Path) -> None:
    db_path = tmp_path / "fund_data.db"
    output_path = tmp_path / "all_sources.xlsx"
    _create_nav_db(
        db_path,
        [
            ("EMAIL01", "邮件基金", "2026-01-01", 1.0, 1.0, "email"),
            ("ZX01", "臻选基金", "2026-01-01", 2.0, 2.0, "zx_excel"),
        ],
    )

    organize_fund_data(db_path, output_path, source="all")

    workbook = load_workbook(output_path, read_only=True, data_only=True)
    assert workbook.sheetnames == ["汇总", "EMAIL01", "ZX01"]
    workbook.close()


def test_email_text_is_exported_as_text_not_excel_formula(tmp_path: Path) -> None:
    db_path = tmp_path / "fund_data.db"
    output_path = tmp_path / "safe_text.xlsx"
    _create_nav_db(
        db_path,
        [
            (
                "SAFE01",
                '=HYPERLINK("https://example.invalid","点击")',
                "2026-01-01",
                1.0,
                1.0,
                "email",
            )
        ],
    )

    organize_fund_data(db_path, output_path, source="email")

    workbook = load_workbook(output_path, read_only=False, data_only=False)
    summary_name = workbook["汇总"]["B2"]
    detail_name = workbook["SAFE01"]["A2"]
    assert summary_name.data_type == "s"
    assert detail_name.data_type == "s"
    assert str(detail_name.value).startswith("'")
    workbook.close()


def test_illegal_xml_characters_are_removed_from_cells_and_sheet_names(tmp_path: Path) -> None:
    db_path = tmp_path / "fund_data.db"
    output_path = tmp_path / "clean_xml.xlsx"
    _create_nav_db(
        db_path,
        [("CTRL\x01/01", "基金\x00名\ufffe称", "2026-01-01", 1.0, 1.0, "email")],
    )

    organize_fund_data(db_path, output_path, source="email")

    workbook = load_workbook(output_path, read_only=True, data_only=True)
    detail_sheet = workbook.worksheets[1]
    assert "\x01" not in detail_sheet.title
    row = next(detail_sheet.iter_rows(min_row=2, values_only=True))
    assert row[0] == "基金名称"
    assert row[1] == "CTRL/01"
    workbook.close()


def test_unparseable_dates_are_skipped_and_reported(tmp_path: Path, capsys) -> None:
    db_path = tmp_path / "fund_data.db"
    output_path = tmp_path / "valid_only.xlsx"
    _create_nav_db(
        db_path,
        [
            ("SAFE01", "有效基金", "2026-01-01", 1.0, 1.0, "email"),
            ("0.9929", "20260629", "基金名称误入日期列", 0.9929, None, "email"),
            ("YEAR", "不完整年份", "2026", 1.0, None, "email"),
            ("SHORT", "缺少补零", "202681", 1.0, None, "email"),
            ("LOOSE", "非规范日期", "2026-8-1", 1.0, None, "email"),
        ],
    )

    organize_fund_data(db_path, output_path, source="email")

    assert "跳过 4 条日期无法识别的记录" in capsys.readouterr().out
    workbook = load_workbook(output_path, read_only=True, data_only=True)
    assert workbook.sheetnames == ["汇总", "SAFE01"]
    assert _sheet_rows(workbook, "汇总")[1][0:3] == ("SAFE01", "有效基金", 1)
    workbook.close()


def test_empty_source_fails_without_creating_a_workbook(tmp_path: Path) -> None:
    db_path = tmp_path / "fund_data.db"
    output_path = tmp_path / "empty.xlsx"
    _create_nav_db(
        db_path,
        [("ZX01", "臻选基金", "2026-01-01", 2.0, 2.0, "zx_excel")],
    )

    with pytest.raises(ValueError, match="没有找到.*净值数据"):
        organize_fund_data(db_path, output_path, source="email")

    assert not output_path.exists()


def test_rejects_database_as_output_without_modifying_it(tmp_path: Path) -> None:
    db_path = tmp_path / "fund_data.db"
    _create_nav_db(
        db_path,
        [("SAFE01", "安全基金", "2026-01-01", 1.0, 1.0, "email")],
    )
    original_bytes = db_path.read_bytes()

    with pytest.raises(ValueError, match="输出文件.*数据库"):
        organize_fund_data(db_path, db_path, source="email")

    assert db_path.read_bytes() == original_bytes
    with closing(sqlite3.connect(db_path)) as conn:
        assert conn.execute("SELECT COUNT(*) FROM fund_nav_data").fetchone()[0] == 1


def test_cli_accepts_database_output_and_source_arguments(tmp_path: Path) -> None:
    db_path = tmp_path / "fund_data.db"
    output_path = tmp_path / "cli_email.xlsx"
    _create_nav_db(
        db_path,
        [("CLI01", "命令行基金", "2026-01-01", 1.0, 1.0, "email")],
    )

    completed = subprocess.run(
        [
            sys.executable,
            str(PROJECT_ROOT / "organize_fund_data.py"),
            "--db",
            str(db_path),
            "--output",
            str(output_path),
            "--source",
            "email",
        ],
        cwd=tmp_path,
        env={**os.environ, "PYTHONIOENCODING": "utf-8"},
        capture_output=True,
        text=True,
        encoding="utf-8",
        timeout=30,
        check=False,
    )

    assert completed.returncode == 0, completed.stderr
    assert output_path.exists()
    assert "1 条记录" in completed.stdout
    assert "1 个基金" in completed.stdout


def test_cli_rejects_database_as_output_without_modifying_it(tmp_path: Path) -> None:
    db_path = tmp_path / "fund_data.db"
    _create_nav_db(
        db_path,
        [("SAFE01", "安全基金", "2026-01-01", 1.0, 1.0, "email")],
    )
    original_bytes = db_path.read_bytes()

    completed = subprocess.run(
        [
            sys.executable,
            str(PROJECT_ROOT / "organize_fund_data.py"),
            "--db",
            str(db_path),
            "--output",
            str(db_path),
            "--source",
            "email",
        ],
        cwd=tmp_path,
        env={**os.environ, "PYTHONIOENCODING": "utf-8"},
        capture_output=True,
        text=True,
        encoding="utf-8",
        timeout=30,
        check=False,
    )

    assert completed.returncode == 1
    assert "输出文件" in completed.stdout
    assert "数据库" in completed.stdout
    assert db_path.read_bytes() == original_bytes
